from __future__ import annotations

from typing import Iterable, List, Dict, Any
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from ..config import config


def autosize(ws) -> None:
    for col_cells in ws.iter_cols(min_row=1):
        length = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in col_cells
        )
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(12, length + 2), 60)


def save_results_workbook(
    results_rows: Iterable[Dict[str, Any]],
    citations_rows: Iterable[Dict[str, Any]],
    inputs_rows: Iterable[Dict[str, Any]] = tuple(),
    filename: str = "result.xlsx"
) -> str:
    wb = Workbook()
    ws_res = wb.active
    ws_res.title = "Results"
    
    # Add question header if available in inputs
    inputs_list = list(inputs_rows)
    question = next((inp.get("value") for inp in inputs_list if inp.get("key") == "Question"), None)
    if question:
        ws_res.append([f"QUESTION: {question}", "", "", ""])
        ws_res.append(["", "", "", ""])  # Empty row for spacing
    
    headers = ["Context", "Evidence", "Answer", "Notes"]
    ws_res.append(headers)
    for row in results_rows:
        ws_res.append([row.get("context"), row.get("evidence"), row.get("answer"), row.get("notes")])
    autosize(ws_res)

    ws_cit = wb.create_sheet("Citations")
    headers_cit = ["Doc Name", "Source URI", "Doc Type", "Statement Type", "Period", "Location", "Quote", "Chunk Id", "Score"]
    ws_cit.append(headers_cit)
    for c in citations_rows:
        period = f"{c.get('year','')}/Q{c.get('quarter','')}" if c.get('year') else ""
        loc = c.get('sheet') and f"{c.get('sheet')}!{c.get('cell_range','')}" or (
            c.get('page') and f"p.{c.get('page')}:{c.get('line_start')}-{c.get('line_end')}" or ""
        )
        ws_cit.append([
            c.get("doc_name"), c.get("source_uri"), c.get("doc_type"), c.get("statement_type"),
            period, loc, c.get("quote"), c.get("chunk_id"), c.get("score")
        ])
    autosize(ws_cit)

    ws_in = wb.create_sheet("Inputs & Assumptions")
    ws_in.append(["Key", "Value"])
    for kv in inputs_rows:
        ws_in.append([kv.get("key"), kv.get("value")])
    autosize(ws_in)

    out_path = os.path.join(config.service.artifacts_dir, filename)
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return out_path
